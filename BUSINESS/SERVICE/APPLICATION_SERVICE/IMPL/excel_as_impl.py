# -*- coding: utf-8 -*-

"""
excel_as_impl.py: The python file dedicated to the implementation class of the Application Service
part dedicated to any need of Excel service by the Application.
"""

__author__ = "Rindra Mbolamananamalala"
__email__ = "rindraibi@gmail.com"

import os
import shutil
import uuid
import decimal

from datetime import datetime

from openpyxl import load_workbook

from UTILS.project_utils import get_project_path

from CONFIGURATIONS.logger import LOGGER
from CONFIGURATIONS.application_properties import get_application_property

from BUSINESS.SERVICE.APPLICATION_SERVICE.INTF.excel_as_intf import ExcelASIntf


def generate_name_for_submition_excel_file(zone_parameter: str) -> str:
    """
    Generating a name for the Excel File for the Submition, while respecting the indications of the Management Rules
    about this name.
    :param zone_parameter: The "Zone" parameter to be included within the file's name
    :return: The name formulated according to the Management Rules corresponding to it
    """
    try:
        date = datetime.today().date()
        time = str(datetime.today().time()).replace(".", ",")  # In order to see the decimals
        unique_id = uuid.uuid4()
        name_generated = str(date) + "_" + str(time) + "_" + zone_parameter + "_" + str(unique_id)
        # making it possible to distinguish the time but at the same time respecting the rules for the name of
        # an Excel file
        return name_generated.replace(":", ".")
    except Exception as exception:
        # At least one error has occurred, therefore, stop the process
        LOGGER.error(
            exception.__class__.__name__ + ": " + str(exception)
            + ". Can't go further with the file name Generation Process. "
        )
        raise


def get_current_shift() -> int:
    """
    In function of the current time and the 3 different Time Stamps, determining the current Shift
    :return: The current Shift
    """
    try:
        # Retrieving all the Time Stamps
        time_stamp_1_min = int(get_application_property("time_stamp_1").split(",")[0])
        time_stamp_1_max = int(get_application_property("time_stamp_1").split(",")[1])
        time_stamp_2_min = int(get_application_property("time_stamp_2").split(",")[0])
        time_stamp_2_max = int(get_application_property("time_stamp_2").split(",")[1])
        time_stamp_3_min = int(get_application_property("time_stamp_3").split(",")[0])
        time_stamp_3_max = int(get_application_property("time_stamp_3").split(",")[1])
        # First, let's get the current Time Stamp
        current_hour = datetime.today().hour
        if time_stamp_1_min <= current_hour < time_stamp_1_max:
            # Shift 1
            return 1
        elif time_stamp_2_min <= current_hour < time_stamp_2_max:
            # Shift 2
            return 2
        elif current_hour >= time_stamp_3_min or current_hour < time_stamp_3_max:
            # Shift 3
            return 3
        else:
            # Should never end up here...
            msg_error = "Impossible to find a Shift for Current Hour : \"" + str(
                datetime.today().time()) + "\""
            LOGGER.error(msg_error)
            raise Exception(msg_error)
    except Exception as exception:
        # At least one error has occurred, therefore, stop the process
        LOGGER.error(
            exception.__class__.__name__ + ": " + str(exception)
            + ". Can't go further with the Shift determination Process. "
        )
        raise


class ExcelASImpl(ExcelASIntf):

    def create_submition_excel_file(self, zone_parameter: str) -> str:
        """
        Creating the Excel file for the "Submition", naming it, storing it in the dedicated folder and then returning
        its whole file path.
        :return: The file path of the Excel file for the "Submition" newly-created.
        """
        try:
            # STEP 1 : Let's have an empty copy of the Submition Excel file template within the folder dedicated to the
            # Excel Files for the actual Submition processes, and name it with the name specifically generated for it
            project_path = get_project_path()
            template_path = project_path + "\\RESOURCES\\EXCEL\\scrap_submition_template.xlsx"
            excel_submition_path = get_application_property("excel_created_folder_path") \
                                   + "\\" \
                                   + generate_name_for_submition_excel_file(zone_parameter) + ".xlsx"
            shutil.copyfile(template_path, excel_submition_path)
            # Excel File for Submition successfully created
            LOGGER.info("The Excel file for the Submition: \"" + excel_submition_path + "\" is successfully created")
            # STEP 2: Returning the whole file Path
            return excel_submition_path
        except Exception as exception:
            # At least one error has occurred, therefore, stop the process
            LOGGER.error(
                exception.__class__.__name__ + ": " + str(exception)
                + ". Can't go further with the file Creation Process. "
            )
            raise

    def insert_categorized_lines(self, team: str
                                        , area_parameter: str
                                        , category: str
                                        , categorized_lines: list
                                        , excel_file_submition_path: str):
        """
        Inserting all the information corresponding to the Categorized Lines within the Excel File for the "Submition".
        :param team: The "Team" parameter selected
        :param area_parameter: The "Area" parameter selected
        :param category: The Category (Material) concerned inside the Excel File
        :param categorized_lines: The categorize lines to be inserted
        :param excel_file_submition_path: The path leading to the Excel File for the Submition process
        :return:
        """
        try:
            # First, let's prepare the Insertion..
            workbook = load_workbook(excel_file_submition_path)
            worksheet = workbook["form_template_srb"]

            # ... feeding static cells...
            worksheet.cell(row=6, column=11).value = area_parameter
            calendar_week = decimal.Decimal(
                    str(datetime.today().isocalendar()[1])
                    + "."
                    + str(datetime.today().isocalendar()[2])
                )
            worksheet.cell(row=8, column=7).value = str(calendar_week) \
                                                    + "." \
                                                    + str(get_current_shift())
            worksheet.cell(row=10, column=7).value = team

            # ... and then, let's specify the starting rows coordinates (within the Excel file) corresponding
            # to each category...
            if category == "Aluminium":
                row_start = 19
            elif category == "Copper":
                row_start = 34
            elif category == "Plastic":
                row_start = 49
            elif category == "Terminal":
                row_start = 64
            elif category == "Harness":
                row_start = 79
            else:
                # Category not (yet) recognized, hope will never end up here...
                msg_error = "Category: \"" + category + "\" not (yet) recognized."
                LOGGER.error(msg_error)
                raise Exception(msg_error)

            # ... starting the determination of the other coordinates for the Insertion...
            counter_line = 1
            for line in categorized_lines:
                if counter_line == 1:
                    row_insertion = row_start
                    col_material_id = 4
                    col_quantity = 9
                elif counter_line == 8:
                    row_insertion = row_start
                    col_material_id = 14
                    col_quantity = 19
                elif counter_line == 15:
                    row_insertion = row_start
                    col_material_id = 24
                    col_quantity = 29
                elif counter_line > 21:
                    # Number of line exceeds the limitation (Automata's WELL)
                    msg_error = "Number of lines: " + str(len(categorized_lines)) + " exceeds the limitation 21."
                    LOGGER.error(msg_error)
                    raise Exception(msg_error)
                else:
                    # Just pass, nothing really special...
                    pass
                # ... actual Insertion...
                worksheet.cell(row=row_insertion, column=col_material_id).value = line.get_material_id()
                worksheet.cell(row=row_insertion, column=col_quantity).value = line.get_quantity()
                counter_line += 1
                row_insertion += 1

            # ... And finally, confirmation of all the actions related to the Insertion.
            workbook.save(excel_file_submition_path)
            workbook.close()
        except Exception as exception:
            # At least one error has occurred, therefore, stop the process
            LOGGER.error(
                exception.__class__.__name__ + ": " + str(exception)
                + ". Can't go further with the Insertion Process. "
            )
            raise

    def print_excel_file(self, excel_file_path: str):
        """
        Printing an Excel file whose path has been specified as an argument with the current (default) printer of
        the Computer.
        :param excel_file_path: The path of the Excel File to be printed
        :return: None
        """
        try:
            os.startfile(excel_file_path, 'print')
            # The Excel file has been successfully printed
            success_msg = "The Excel file : \"" + excel_file_path + "\"" + "has been successfully printed."
            LOGGER.info(success_msg)
        except Exception as exception:
            # At least one error has occurred, therefore, stop the process
            LOGGER.error(
                exception.__class__.__name__ + ": " + str(exception)
                + ". Can't go further with the Printing Process. "
            )
            raise
